"""Excel ingestion boundary for Phase 2C.

This module performs deterministic Excel ingestion using openpyxl. It reads the
first worksheet, normalizes headers, validates required columns, records
warnings/errors, and persists rows to existing SQLite tables. The ingestion
boundary does not implement business logic, classification, or UI prompts.
"""

from __future__ import annotations

import json
import logging
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable, Mapping, Protocol, Sequence

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class DatabaseLike(Protocol):
    """Database contract required by the ingestion boundary."""

    def connect(self) -> sqlite3.Connection: ...


class RuntimePathsLike(Protocol):
    """Runtime paths contract (used for logging context only)."""

    run_id: str


@dataclass(frozen=True)
class ValidationIssue:
    """Structured validation issue captured during ingestion."""

    row_number: int | None
    column_name: str | None
    error_type: str
    message: str
    severity: str


class IngestionFatalError(RuntimeError):
    """Fatal ingestion error that stops processing for the current file."""

    def __init__(self, message: str, issues: Sequence[ValidationIssue]) -> None:
        super().__init__(message)
        self.issues = tuple(issues)


@dataclass(frozen=True)
class IngestionStats:
    """Per-source ingestion results."""

    source_name: str
    file_path: Path
    rows_seen: int
    rows_inserted: int
    warnings: tuple[ValidationIssue, ...]


@dataclass(frozen=True)
class IngestionSummary:
    """Aggregate ingestion results across all sources."""

    results: tuple[IngestionStats, ...]
    warnings: tuple[ValidationIssue, ...]
    has_warnings: bool


@dataclass(frozen=True)
class SheetSpec:
    """Specification for ingesting a single Excel source."""

    source_name: str
    required_columns: tuple[str, ...]
    key_columns: tuple[str, ...]
    warning_columns: tuple[str, ...]
    table_name: str
    row_mapper: Callable[[Mapping[str, object]], Mapping[str, object]]


TABLE_VESSEL = "vessel"
TABLE_VESSEL_INVENTORY = "vessel_inventory_row"
TABLE_IC_INVENTORY = "ic_inventory_row"


IC_REQUIRED = (
    "item",
    "itmdesc",
    "plinid",
    "itmclss",
    "upccode",
    "edition",
    "currdate",
)
VESSEL_INDEX_REQUIRED = (
    "shipid",
    "shipname",
    "custno",
    "imono",
    "shipstat",
    "email",
    "note1",
    "note2",
    "note3",
)
VESSEL_INVENTORY_REQUIRED = (
    "shipid",
    "shipname",
    "custno",
    "item",
    "edition",
    "storeedt",
    "descrip",
)


IC_SPEC = SheetSpec(
    source_name="safe_ic_inventory",
    required_columns=IC_REQUIRED,
    key_columns=("item",),
    warning_columns=(),
    table_name=TABLE_IC_INVENTORY,
    row_mapper=lambda row: {
        "item": row.get("item"),
        "current_edition": row.get("edition"),
        "description": row.get("itmdesc"),
        "current_date": row.get("currdate"),
    },
)

VESSEL_INDEX_SPEC = SheetSpec(
    source_name="safe_vessels_index",
    required_columns=VESSEL_INDEX_REQUIRED,
    key_columns=("shipid",),
    warning_columns=("email",),
    table_name=TABLE_VESSEL,
    row_mapper=lambda row: {
        "ship_id": row.get("shipid"),
        "ship_name": row.get("shipname"),
        "customer_no": row.get("custno"),
        "imo_no": row.get("imono"),
        "ship_status": row.get("shipstat"),
        "ship_email": row.get("email"),
        "office_email": None,
        "ams": None,
    },
)

VESSEL_INVENTORY_SPEC = SheetSpec(
    source_name="safe_vessels_inventory",
    required_columns=VESSEL_INVENTORY_REQUIRED,
    key_columns=("shipid", "item"),
    warning_columns=(),
    table_name=TABLE_VESSEL_INVENTORY,
    row_mapper=lambda row: {
        "ship_id": row.get("shipid"),
        "item": row.get("item"),
        "onboard_edition": row.get("edition"),
        "store_edition": row.get("storeedt"),
        "description": row.get("descrip"),
    },
)


def ingest_excel_files(
    *,
    ic_inventory_path: Path | str,
    vessels_index_path: Path | str,
    vessels_inventory_path: Path | str,
    db: DatabaseLike,
    paths: RuntimePathsLike,
) -> IngestionSummary:
    """Ingest the three required Excel workbooks into SQLite.

    Raises:
        IngestionFatalError: If a fatal schema or file error occurs.
    """

    results = (
        _ingest_single_file(Path(ic_inventory_path), IC_SPEC, db, paths),
        _ingest_single_file(Path(vessels_index_path), VESSEL_INDEX_SPEC, db, paths),
        _ingest_single_file(Path(vessels_inventory_path), VESSEL_INVENTORY_SPEC, db, paths),
    )
    warnings = tuple(issue for result in results for issue in result.warnings)
    return IngestionSummary(results=results, warnings=warnings, has_warnings=bool(warnings))


def _ingest_single_file(
    file_path: Path,
    spec: SheetSpec,
    db: DatabaseLike,
    paths: RuntimePathsLike,
) -> IngestionStats:
    """Ingest a single Excel workbook into its target table."""

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - error surface only
        issue = ValidationIssue(
            row_number=None,
            column_name=None,
            error_type="unreadable_file",
            message=f"{spec.source_name}: failed to read {file_path}: {exc}",
            severity="fatal",
        )
        _persist_validation_issues(db, [issue])
        raise IngestionFatalError(issue.message, [issue]) from exc

    try:
        if not workbook.worksheets:
            issue = ValidationIssue(
                row_number=None,
                column_name=None,
                error_type="empty_worksheet",
                message=f"{spec.source_name}: workbook has no worksheets: {file_path}",
                severity="fatal",
            )
            _persist_validation_issues(db, [issue])
            raise IngestionFatalError(issue.message, [issue])

        worksheet = workbook.worksheets[0]
        header_values = _read_header_row(worksheet)
        if header_values is None:
            issue = ValidationIssue(
                row_number=None,
                column_name=None,
                error_type="empty_worksheet",
                message=f"{spec.source_name}: worksheet is empty in {file_path}",
                severity="fatal",
            )
            _persist_validation_issues(db, [issue])
            raise IngestionFatalError(issue.message, [issue])

        if not any(not _is_blank(cell) for cell in header_values):
            issue = ValidationIssue(
                row_number=1,
                column_name=None,
                error_type="missing_header",
                message=f"{spec.source_name}: missing header row in {file_path}",
                severity="fatal",
            )
            _persist_validation_issues(db, [issue])
            raise IngestionFatalError(issue.message, [issue])

        header_map, header_warnings = _normalize_headers(header_values, spec)
        missing_columns = sorted(set(spec.required_columns) - set(header_map))
        issues: list[ValidationIssue] = []
        if header_warnings:
            issues.extend(header_warnings)

        if missing_columns:
            fatal_issues = [
                ValidationIssue(
                    row_number=1,
                    column_name=column_name,
                    error_type="missing_required_column",
                    message=f"{spec.source_name}: missing required column '{column_name}'",
                    severity="fatal",
                )
                for column_name in missing_columns
            ]
            issues.extend(fatal_issues)
            _log_warnings(issues, paths.run_id)
            _persist_validation_issues(db, issues)
            raise IngestionFatalError(
                f"{spec.source_name}: missing required columns {missing_columns}", issues
            )

        rows_seen = 0
        rows_inserted = 0
        row_payloads: list[tuple[int, str]] = []
        table_payloads: list[Mapping[str, object]] = []
        row_issues: list[ValidationIssue] = []

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            rows_seen += 1
            normalized_row = _extract_row(row, header_map)
            if _is_empty_row(normalized_row):
                row_issues.append(
                    ValidationIssue(
                        row_number=row_number,
                        column_name=None,
                        error_type="empty_row",
                        message=f"{spec.source_name}: empty row {row_number}",
                        severity="warning",
                    )
                )
                continue

            missing_keys = [
                key for key in spec.key_columns if _is_blank(normalized_row.get(key))
            ]
            if missing_keys:
                for key in missing_keys:
                    row_issues.append(
                        ValidationIssue(
                            row_number=row_number,
                            column_name=key,
                            error_type="missing_key_field",
                            message=(
                                f"{spec.source_name}: missing key field '{key}' "
                                f"on row {row_number}"
                            ),
                            severity="warning",
                        )
                    )
                continue

            for warn_column in spec.warning_columns:
                if _is_blank(normalized_row.get(warn_column)):
                    row_issues.append(
                        ValidationIssue(
                            row_number=row_number,
                            column_name=warn_column,
                            error_type="missing_optional_field",
                            message=(
                                f"{spec.source_name}: missing optional field "
                                f"'{warn_column}' on row {row_number}"
                            ),
                            severity="warning",
                        )
                    )

            row_payloads.append((row_number, _serialize_row(spec.source_name, normalized_row)))
            table_payloads.append(spec.row_mapper(normalized_row))
            rows_inserted += 1

        all_issues = [*issues, *row_issues]
        if all_issues:
            _log_warnings(all_issues, paths.run_id)

        with db.connect() as conn:
            with conn:
                if row_payloads:
                    conn.executemany(
                        "INSERT INTO raw_excel_rows (row_number, row_json) VALUES (?, ?);",
                        row_payloads,
                    )
                if table_payloads:
                    _insert_table_rows(conn, spec.table_name, table_payloads)
                if all_issues:
                    _insert_validation_issues(conn, all_issues)

        warnings = tuple(issue for issue in all_issues if issue.severity == "warning")
        return IngestionStats(
            source_name=spec.source_name,
            file_path=file_path,
            rows_seen=rows_seen,
            rows_inserted=rows_inserted,
            warnings=warnings,
        )
    finally:
        workbook.close()


def _read_header_row(worksheet) -> Sequence[object] | None:
    """Return the first row of values or None if the worksheet is empty."""

    rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    header_values = next(rows, None)
    return header_values


def _normalize_headers(
    header_values: Sequence[object],
    spec: SheetSpec,
) -> tuple[dict[str, int], list[ValidationIssue]]:
    """Normalize headers, handling duplicates with warnings."""

    header_map: dict[str, int] = {}
    warnings: list[ValidationIssue] = []
    for idx, cell in enumerate(header_values):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        if normalized in header_map:
            warnings.append(
                ValidationIssue(
                    row_number=1,
                    column_name=normalized,
                    error_type="duplicate_header",
                    message=(
                        f"{spec.source_name}: duplicate header '{normalized}' "
                        "uses first occurrence"
                    ),
                    severity="warning",
                )
            )
            continue
        header_map[normalized] = idx
    return header_map, warnings


def _normalize_header(value: object) -> str:
    """Normalize header names using trim and casefold."""

    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip().casefold()
    return str(value).strip().casefold()


def _normalize_cell(value: object) -> object:
    """Normalize cell values without applying business logic."""

    if isinstance(value, str):
        return value.strip()
    return value


def _extract_row(row: Sequence[object], header_map: Mapping[str, int]) -> dict[str, object]:
    """Build a normalized row mapping using the normalized headers."""

    normalized: dict[str, object] = {}
    for header, index in header_map.items():
        value = row[index] if index < len(row) else None
        normalized[header] = _normalize_cell(value)
    return normalized


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _is_empty_row(row: Mapping[str, object]) -> bool:
    return all(_is_blank(value) for value in row.values())


def _serialize_row(source_name: str, row: Mapping[str, object]) -> str:
    """Serialize a row for raw storage with a source label."""

    payload = {"source": source_name, "row": {k: _json_safe(v) for k, v in row.items()}}
    return json.dumps(payload, ensure_ascii=True)


def _json_safe(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _adapt_sql_value(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _insert_table_rows(
    conn: sqlite3.Connection,
    table_name: str,
    rows: Iterable[Mapping[str, object]],
) -> None:
    """Insert mapped rows into the target table."""

    rows_list = list(rows)
    if not rows_list:
        return
    columns = list(rows_list[0].keys())
    placeholders = ", ".join("?" for _ in columns)
    column_list = ", ".join(columns)
    sql = f"INSERT INTO {table_name} ({column_list}) VALUES ({placeholders});"
    conn.executemany(
        sql,
        ([_adapt_sql_value(row.get(col)) for col in columns] for row in rows_list),
    )


def _persist_validation_issues(db: DatabaseLike, issues: Sequence[ValidationIssue]) -> None:
    """Persist validation issues in their own transaction."""

    if not issues:
        return
    with db.connect() as conn:
        with conn:
            _insert_validation_issues(conn, issues)


def _insert_validation_issues(
    conn: sqlite3.Connection, issues: Sequence[ValidationIssue]
) -> None:
    conn.executemany(
        """
        INSERT INTO validation_errors (
            row_number,
            column_name,
            error_type,
            message,
            severity
        )
        VALUES (?, ?, ?, ?, ?);
        """,
        [
            (
                issue.row_number,
                issue.column_name,
                issue.error_type,
                issue.message,
                issue.severity,
            )
            for issue in issues
        ],
    )


def _log_warnings(issues: Sequence[ValidationIssue], run_id: str) -> None:
    for issue in issues:
        if issue.severity == "warning":
            logger.warning("[%s] %s", run_id, issue.message)
